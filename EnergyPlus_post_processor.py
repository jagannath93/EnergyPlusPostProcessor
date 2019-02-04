""" 
Post-processor for the scale reduction of EnergyPlus simulation output(CSV).
"""

# Imports
from collections import OrderedDict
import itertools
import string
import math
import openpyxl
import csv
import os


# Globals
# Considering 24hr day as 7:00 AM to 6:00 AM
groups_info = {
    'Bed-Room': {'part-1' : {'3hr': [5,6,7]}},

    'Living-Room': {'part-1': {'3hr':[0]},
                    'part-2': {'3hr':[3,4]}},

    'Kitchen': {'part-1': {'3hr': [0]},
                'part-2': {'1hr': [6,5,4]}, # 11,12,13 -> 6,5,4(indices)
                'part-3': {'3hr': [4]}},

    'Half-Day-Blocks': {'part-1': {'3hr': [0,1,2,3]},
                    'part-2': {'3hr': [4,5,6,7]}}
}

"""
wall_info = {
    'wall_1': ['C', 'D', 'G', 'H'],
    'wall_2': ['I', 'J', 'M', 'N'],
    'wall_3': ['O', 'P', 'S', 'T'],
    'wall_4': ['U', 'V', 'Y', 'Z'],
    'wall_5': ['AA', 'AB', 'AE', 'AF'],
    'wall_6': ['AG', 'AH', 'AK', 'AL'],
    'wall_7': ['AM', 'AN', 'AQ', 'AR'],
    'wall_8': ['AS', 'AT', 'AW', 'AX'],
    'wall_9': ['AY', 'AZ', 'BC', 'BD'],
    'wall_10': ['BE', 'BF', 'BI', 'BJ'],
    'wall_11': ['BK', 'BL', 'BO', 'BP'],
    'wall_12': ['BQ', 'BR', 'BU', 'BV']
}

room_info = {
    'room-1': ['CJ', 'DP'],
    'room-2': ['CK', 'DQ'],
    'room-3': ['CL', 'DR'],
    'room-4': ['CM', 'DS'],
    'room-5': ['CN', 'DU'],
    'room-6': ['CU', 'DV'],
    'room-7': ['CP', 'DW'],
    'room-8': ['CQ', 'DX'],
    'room-9': ['CR', 'DZ'],
    'room-10': ['CS', 'EA']
}

room_name_no_map = {
  'Room-1': 'Room-2465',
  'Room-2': 'Room-2472',
  'Room-3': 'Room-2479',
  'Room-4': 'Room-2486',
  'Room-5': 'Room-2500',
  'Room-6': 'Room-2542',
  'Room-7': 'Room-2535',
  'Room-8': 'Room-2507',
  'Room-9': 'Room-2514',
  'Room-10': 'Room-2521',
}
"""

# Definitions
def read_csvfile(filepath):
    with open(filepath, 'rb') as f:
        reader_data = csv.reader(f, delimiter=',')
        data = list(reader_data)
        col_names = data[0]
        data = data[1:]
        data = data[6:] + data[:6]

        # Generate letters for columns
        l1 = list(string.ascii_uppercase)
        l2 = map(lambda x: ''.join(x), itertools.product(l1, repeat=2))
        l = l1 + l2

        # Map column Name, Column Letter and Serial Number
        for counter, (col, name) in enumerate(zip(l, col_names)):
            col_map[col] = counter
            col_name_map[col] = name
    return data

def create_data_bins(data):
    data_bins['365day'] = data
    data_bins['24hr'] = OrderedDict()
    counter = 1
    for i in xrange(0, 8759, 24): # 8760 - 1
        day = 'day_'+str(counter)
        data_bins['24hr'][day] = {}

        # Generate 1hr bins
        _1hr_bins = data_bins['24hr'][day]['1hr'] = data[i:i+24]

        # Generate 3hr bins.
        _3hr_bins = data_bins['24hr'][day]['3hr'] = []
        for j in xrange(0, len(_1hr_bins), 3):
            _3hr_bins.append(_1hr_bins[j:j+3])

        counter += 1

    data_bins['summer'] = data[3216:3960]   # Data division for 'Summer' Bin
    data_bins['winter'] = data[8015:8759]   # Data division for 'Winter' Bin

def create_data_groups():
    group_data = data_bins['groups'] = {}

    for group, info in groups_info.items():
        tmp = []
        for item in info.values():
            for key, values in item.items():
                for value in values:
                    for i in xrange(1, 365+1):
                        if key == '3hr':
                            tmp.extend(data_bins['24hr']['day_'+str(i)]['3hr'][value])
                        elif key == '1hr':
                            tmp.append(data_bins['24hr']['day_'+str(i)]['1hr'][value])

        group_data[group] = tmp


# Re-check the Data
def recheck_data():
    #print "\n\n---------- Re-Checking the Input ----------", "green"
    print "\nKeys in Data Bin: " + ",".join(data_bins.keys())
    print "\n365day Bin Size: " + str(len(data_bins['365day']))
    print "\n24hr Bin Size: " + str(len(data_bins['24hr']))
    print "\nSummer: " + str(len(data_bins['summer']))
    print "\nWinter: " + str(len(data_bins['winter']))

    print "\n1hr Sub-Bin Size: " + str(len(data_bins['24hr']['day_1']['1hr']))
    print "\n3hr Sub-Bin Size: " + str(len(data_bins['24hr']['day_1']['3hr']))

    print "\nDifferent Groups: " + ",".join(data_bins['groups'].keys())
    print "\nBed-Room Group Size: " + str(len(data_bins['groups']['Bed-Room']))
    print "\nBed-Room Data Item Size: " + str(len(data_bins['groups']['Bed-Room'][0]))
    print "\nLiving-Room Group Size: " + str(len(data_bins['groups']['Living-Room']))
    print "\nLiving-Room Data Item Size: " + str(len(data_bins['groups']['Living-Room'][0]))
    print "\nKitchen Group Size: " + str(len(data_bins['groups']['Kitchen']))
    print "\nKitchen Data Item Size: " + str(len(data_bins['groups']['Kitchen'][0]))
    print "\nHalf-Day-Blocks Group Size: " + str(len(data_bins['groups']['Half-Day-Blocks']))
    print "\nHalf-Day-Blocks Data Items Size: " + str(len(data_bins['groups']['Half-Day-Blocks'][0]))


		
# Main Processing Definitions	

# Create Excel Workbook
def create_excel_workbook(output_batch_filename):
    wb = openpyxl.Workbook()
    wb.save(output_batch_filename)
		#return wb
		
# Read Input 'Wall' and 'Room' Information.
def read_input_info(input_info_file_path):
	wb = openpyxl.load_workbook(input_info_file_path)
	
	ws = wb.get_sheet_by_name("walls")
	rows = ws.rows
	wall_param_names.extend(map(lambda x: str(x.value), rows[0][1:]))
	for row in rows[1:]:
		wall_name = str(row[0].value)
		wall_data = map(lambda x: x.value, row[1:])
		wall_info[wall_name] = wall_data
		
	ws = wb.get_sheet_by_name("rooms_type1")
	rows = ws.rows
	room_param_type1_names.extend(map(lambda x: str(x.value), rows[0][1:]))
	for row in rows[1:]:
		room_name = str(row[0].value)
		room_data = map(lambda x: x.value, row[1:])
		room_info_type1[room_name] = room_data
		
	ws = wb.get_sheet_by_name("rooms_type2")
	rows = ws.rows
	room_param_type2_names.extend(map(lambda x: str(x.value), rows[0][1:]))
	for row in rows[1:]:
		room_name = str(row[0].value)
		room_data = map(lambda x: x.value, row[1:])
		room_info_type2[room_name] = room_data

# Task Definitions
def do_task1():
    output_bins
    output_bins['365day'] = {}
    days = output_bins['365day']['days'] = OrderedDict()
    for i in xrange(0, 365):
        day = days['day_'+str(i+1)] = {}
        tmp = []
        rows_list = data_bins['24hr']['day_'+str(i+1)]['1hr']
        tmp.append(get_max('B', rows_list))
        tmp.append(get_min('B', rows_list))
        tmp.append(get_mean('B', rows_list))
        tmp.append(get_range('B', rows_list))
        day['B'] = tmp
    #print output_bins['365day']['days']['day_1']['B']
    return output_bins

def prepare_task1(wb):
	sheet = "Task-1"
	ws = wb.get_active_sheet()
	ws.title = sheet
	#ws.merge_cells("A1:E1")
	#ws['A1'] = 'Outside Drybulb Temperature'
	col_names = ['Filename_Day', 'Maximum', 'Minimum', 'Mean', 'Range']
	ws.append(col_names)
	ws.freeze_panes = 'A2'
	return wb

def export_task1(wb, filename):
	sheet = "Task-1"
	ws = wb.get_sheet_by_name(sheet)

	for i, (key, value) in enumerate(output_bins['365day']['days'].items()):
			data = value['B']
			_data = [filename+'_day'+str(i+1), data[0], data[1], data[2], data[3]]
			ws.append(_data)
	return wb

	
def do_task2():
    block_s = output_bins['summer'] = {}
    block_w = output_bins['winter'] = {}
    days_s = block_s['days'] = OrderedDict()
    days_w = block_w['days'] = OrderedDict()

    counter = 1
    for i in xrange(0, 744, 24):
        day_s = days_s['day_'+str(counter)] = {}
        day_w = days_w['day_'+str(counter)] = {}
        walls_s = day_s['walls'] = OrderedDict()
        walls_w = day_w['walls'] = OrderedDict()
        for name, info in wall_info.items():
            wall_s = walls_s[name] = []
            wall_w = walls_w[name] = []
            for col in info:
                tmp_s = []
                tmp_w = []

                # Calculations for 'Summer' bin
                rows_list_s = data_bins['summer'][i:i+24]
                tmp_s.append(get_max(col, rows_list_s))
                tmp_s.append(get_min(col, rows_list_s))
                tmp_s.append(get_mean(col, rows_list_s))

                # Calculations for 'Winter' bin
                rows_list_w = data_bins['winter'][i:i+24]
                tmp_w.append(get_max(col, rows_list_w))
                tmp_w.append(get_min(col, rows_list_w))
                tmp_w.append(get_mean(col, rows_list_w))

                wall_s.append(tmp_s)
                wall_w.append(tmp_w)
        counter += 1
    return output_bins

def prepare_task2(wb):
	sheet = "Task-2"
	ws = wb.create_sheet()
	ws.title = sheet
	col_names = ['Day']
	for wall_name, info in wall_info.items():
		for col, col_name in zip(info, wall_param_names):
			col_names.extend([wall_name+'_'+col_name+'_Maximum', wall_name+'_'+col_name+'_Minimum', wall_name+'_'+col_name+'_Mean'])
	ws.append(col_names)
	ws.freeze_panes = 'A2'
	return wb

def export_task2(wb, filename):
	sheet = "Task-2"
	ws = wb.get_sheet_by_name(sheet)

	for i in xrange(0, 31):
			data_s = output_bins['summer']['days']['day_'+str(i+1)]['walls']

			_data_s = [filename+'_day'+str(i+1)+'_summer']
			for value in data_s.values():
					for x in value:
							_data_s.extend(x)

			ws.append(_data_s)

	for i in xrange(0, 31):
			data_w = output_bins['winter']['days']['day_'+str(i+1)]['walls']

			_data_w = [filename+'_day'+str(i+1)+'_winter']
			for value in data_w.values():
					for x in value:
							_data_w.extend(x)

			ws.append(_data_w)
	return wb

	
def do_task3():
	rooms = output_bins['365day']['rooms'] = OrderedDict()
	rooms_s = output_bins['summer']['rooms'] = OrderedDict()
	rooms_w = output_bins['winter']['rooms'] = OrderedDict()

	for name, info in room_info_type2.items():
		rooms[name] = []
		rooms_s[name] = []
		rooms_w[name] = []
		for col in info:
			# Sum across 365day in 8 time-blocks
			sum_365day = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
			for i in xrange(0, 365):
				hr_blocks = data_bins['24hr']['day_'+str(i+1)]['1hr']
				tmp = []
				for j in xrange(0, 24, 3):
						tmp.append(sum(map(lambda x: float(x[col_map[col]]), hr_blocks[j:j+3])))
				sum_365day = [sum(x) for x in zip(sum_365day, tmp)]

			# Sum across the summer and winter in 8 time-blocks
			sum_summer = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]
			sum_winter = [0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0]

			for i in xrange(0, 31):
				s_hr_blocks = data_bins['summer']
				w_hr_blocks = data_bins['winter']
				tmp_s = []
				tmp_w = []
				for j in xrange(0, 24, 3):
					tmp_s.append(sum(map(lambda x: float(x[col_map[col]]), s_hr_blocks[j:j+3])))
					tmp_w.append(sum(map(lambda x: float(x[col_map[col]]), w_hr_blocks[j:j+3])))
				sum_summer = [sum(x) for x in zip(sum_summer, tmp_s)]
				sum_winter = [sum(x) for x in zip(sum_winter, tmp_w)]

			rooms[name].extend(sum_365day)
			rooms_s[name].extend(sum_summer)
			rooms_w[name].extend(sum_winter)

			#print sum_365day, sum_summer, sum_winter
			#print len(sum_365day), len(sum_summer), len(sum_winter)

	return output_bins

def prepare_task3(wb):
	sheet = "Task-3"
	ws = wb.create_sheet()
	ws.title = sheet
	col_names = ['File']
	for name, info in room_info_type2.items():
		for col, col_name in zip(info, room_param_type2_names):
			prefix = name +'_'+ col_name
			col_names.extend([prefix+'_block-1_365', prefix+'_block-2_365', prefix+'_block-3_365', prefix+'_block-4_365', prefix+'_block-5_365', prefix+'_block-6_365', prefix+'_block-7_365', prefix+'_block-8_365'])
			col_names.extend([prefix+'_block-1_summer', prefix+'_block-2_summer', prefix+'_block-3_summer', prefix+'_block-4_summer', prefix+'_block-5_summer', prefix+'_block-6_summer', prefix+'_block-7_summer', prefix+'_block-8_summer'])
			col_names.extend([prefix+'_block-1_winter', prefix+'_block-2_winter', prefix+'_block-3_winter', prefix+'_block-4_winter',prefix+'_block-5_winter', prefix+'_block-6_winter', prefix+'_block-7_winter', prefix+'_block-8_winter'])

		#col_names = col_names + (base_titles * len(room_param_type2_names))

	ws.append(col_names)
	ws.freeze_panes = 'A2'
	return wb

def export_task3(wb, filename):
	sheet = "Task-3"
	ws = wb.get_sheet_by_name(sheet)

	_data = [filename]
	for name, info in room_info_type2.items():
		_data.extend(output_bins['365day']['rooms'][name])
		_data.extend(output_bins['summer']['rooms'][name])
		_data.extend(output_bins['winter']['rooms'][name])

	ws.append(_data)
	return wb


def do_task4():
	output_bins
	for i in xrange(0, 365):
		days = output_bins['365day']['days']
		rooms = days['day_'+str(i+1)]['rooms'] = OrderedDict()
		for name, info in room_info_type1.items():
			room = rooms[name] = []
			for col in info:
				#col = info[1]
				rows_list = data_bins['24hr']['day_'+str(i+1)]['1hr']
				room.append(get_max(col, rows_list))
				room.append(get_min(col, rows_list))
				room.append(get_mean(col, rows_list))

				# Calculating 'Damping Factor'(df)
				inside_temp_range = get_range(col, rows_list)
				#print output_bins['365day']['days']['day_1']['rooms'].keys()
				outside_temp_range = output_bins['365day']['days']['day_'+str(i+1)]['B'][3]
				df = ((outside_temp_range - inside_temp_range)/outside_temp_range)*100

				room.append(inside_temp_range)
				room.append(df)
				#print room
	return output_bins

def prepare_task4(wb):
	sheet = "Task-4"
	ws = wb.create_sheet()
	ws.title = sheet
	col_names = ['Day']
	for room_name, info in room_info_type1.items():
		for col, col_name in zip(info, room_param_type1_names):
			col_names.extend([room_name+'_'+col_name+'_Maximum', room_name+'_'+col_name+'_Minimum', room_name+'_'+col_name+'_Mean', room_name+'_'+col_name+'_Range', room_name+'_'+col_name+'_Damping-Factor'])
	ws.append(col_names)
	ws.freeze_panes = 'A2'
	return wb

def export_task4(wb, filename):
	sheet = "Task-4"
	ws = wb.get_sheet_by_name(sheet)

	for i in xrange(0, 365):
			_data = [filename+"_day"+str(i+1)]
			for value in output_bins['365day']['days']['day_'+str(i+1)]['rooms'].values():
				for item in value:
					_data.append(item)
			ws.append(_data)
	return wb


def do_task5():
	corr = output_bins['correlations'] = {}
	corr['rooms'] = OrderedDict()
	tmp = {}
	for name, info in room_info_type1.items():
		output_bins['correlations']['rooms'][name] = []
		for col in info:
			r2_max = corr_util('B', 'max', name, 'max')
			r2_min = corr_util('B', 'min', name, 'min')
			r2_mean = corr_util('B', 'mean', name, 'mean')
			output_bins['correlations']['rooms'][name].extend([r2_max, r2_min, r2_mean])
	return output_bins

def prepare_task5(wb):
	sheet = "Task-5"
	ws = wb.create_sheet()
	ws.title = sheet
	col_names = ['File']
	for room_name, info in room_info_type1.items():
		for col, col_name in zip(info, room_param_type1_names):
			col_names.extend([room_name+'_'+col_name+'_and_T(out)_r^2_Maximum', room_name+'_'+col_name+'_and_T(out)_r^2_Minimum', room_name+'_'+col_name+'_and_T(out)_r^2_Mean'])
	ws.append(col_names)
	ws.freeze_panes = 'A2'
	return wb

def export_task5(wb, filename):
	sheet = "Task-5"
	ws = wb.get_sheet_by_name(sheet)
	_data = [filename]
	for value in output_bins['correlations']['rooms'].values():
		for item in value:
			_data.append(item)
	ws.append(_data)
	return wb
	
	
"""
def do_task6():
	data = data_bins['365day']
	tmp1 = []
	for i in xrange(0, 8759):
			item = data[i]
			_val = item[col_map['EG']]
			if _val is not '':
					tmp1.append(float(_val))

	for room_name in room_info.keys():
			mean_temps = []
			for i in xrange(0, 365):
					mean_temps.append(output_bins['365day']['days']['day_'+str(i+1)]['rooms'][room_name][2])

			corr = get_correlation(tmp1, mean_temps)
			output_bins['correlations']['rooms'][room_name].append(corr)
	return output_bins
"""
	
def do_task7(day_no):
    #for i in [1,80, 141, 172, 265, 355]:  # Specific Days
    days = output_bins['365day']['days']
    rooms = days['day_'+str(day_no)]['rooms'] = OrderedDict()
    for name, info in room_info.items():
      room = rooms[name] = []
      col = info[0]
      rows_list = data_bins['24hr']['day_'+str(day_no)]['1hr']
      room.append(get_max(col, rows_list))
      room.append(get_min(col, rows_list))
    return output_bins

def prepare_task7(room_name_old, room_name_new, day_no):
    sheet = "Task-7"
    output_name = room_name_new +"_"+ day_name_no_map[day_no]
    output_name = output_name + ".xlsx"
    wb = openpyxl.Workbook()
    ws = wb.create_sheet()
    ws.title = sheet
    wb.save(output_name)


day_name_no_map = {
  1: 'Jan-1st',
  80: 'Mar-21st',
  141: 'May-21st',
  172: 'June-21st',
  266: 'Sep-23rd',
  355: 'Dec-21st'
}

def export_task7(room_name_old, room_name_new, day_no):
    sheet = "Task-7"
    output_name = room_name_new +"_"+ day_name_no_map[day_no]
    output_name = output_name + ".xlsx"
    wb = openpyxl.load_workbook(output_name)
    ws = wb.get_sheet_by_name(sheet)

    #for i in xrange(0, 365):
    _data = output_bins['365day']['days']['day_'+str(day_no)]['rooms'][room_name_old]
    #_data.extend([value[0], value[1])
    ws.append(_data)
    wb.save(output_name)

###### Temporary Tasks  #######
"""
def do_task8(room_name_old, day_no):
    block_s = output_bins['summer'] = {}
    block_w = output_bins['winter'] = {}
    days_s = block_s['days'] = OrderedDict()
    days_w = block_w['days'] = OrderedDict()

    day_s = days_s['day_'+str(day_no)] = {}
    day_w = days_w['day_'+str(day_no)] = {}
    walls_s = day_s['walls'] = OrderedDict()
    walls_w = day_w['walls'] = OrderedDict()
    for name, info in wall_info.items():
      wall_s = walls_s[name] = []
      wall_w = walls_w[name] = []
      for col in info:
        tmp_s = []
        tmp_w = []

        # Calculations for 'Summer' bin
        rows_list_s = data_bins['summer'][i:i+24]
        tmp_s.append(get_max(col, rows_list_s))
        tmp_s.append(get_min(col, rows_list_s))
        tmp_s.append(get_mean(col, rows_list_s))

         # Calculations for 'Winter' bin
         rows_list_w = data_bins['winter'][i:i+24]
         tmp_w.append(get_max(col, rows_list_w))
         tmp_w.append(get_min(col, rows_list_w))
         tmp_w.append(get_mean(col, rows_list_w))

         wall_s.append(tmp_s)
         wall_w.append(tmp_w)            

    return output_bins
"""  
#############################

# Utility Definitions
def corr_util(col, type_1, room, type_2):
    _indexes = {'max': 0, 'min': 1, 'mean': 2, 'range': 3}
    values_x = []
    values_y = []
    for i in xrange(0, 364):
        x = output_bins['365day']['days']['day_'+str(i+1)][col][_indexes[type_1]]
        y = output_bins['365day']['days']['day_'+str(i+1)]['rooms'][room][_indexes[type_2]]
        values_x.append(float(x))
        values_y.append(float(y))
    return get_correlation(values_x, values_y)

def get_correlation(values_x, values_y):
    _mean_x = sum(values_x)/float(len(values_x))
    _mean_y = sum(values_y)/float(len(values_y))
    n = len(values_x) # equals to len(values_y) also

    _sum1 = 0.0
    _sum2_x = 0.0
    _sum2_y = 0.0
    for j in xrange(0, n):
        diff_x = values_x[j] - _mean_x
        diff_y = values_y[j] - _mean_y
        _sum1 += (diff_x * diff_y)
        _sum2_x += math.pow(diff_x, 2)
        _sum2_y += math.pow(diff_y, 2)

    _covariance = _sum1/float((n-1))
    SD_x = math.sqrt(_sum2_x/float((n-1)))
    SD_y = math.sqrt(_sum2_y/float((n-1)))
    # TODO: Handle negative square roots here
    _correlation = _covariance/(SD_x * SD_y)
    return math.pow(_correlation, 2)

def get_max(col_letter, rows_list):
    col_index = col_map[col_letter]
    tmp = []
    for row in rows_list:
        tmp.append(float(row[col_index]))
    return max(tmp)

def get_min(col_letter, rows_list):
    col_index = col_map[col_letter]
    tmp = []
    for row in rows_list:
        tmp.append(float(row[col_index]))
    return min(tmp)

def get_mean(col_letter, rows_list):
    col_index = col_map[col_letter]
    tmp = []
    for row in rows_list:
        tmp.append(float(row[col_index]))
    return sum(tmp)/float(len(tmp))

def get_range(col_letter, rows_list):
    return get_max(col_letter, rows_list) - get_min(col_letter, rows_list)

#def get_damping_factor(col_letter, rows_list):
#    return get_range()

def check_output():
    #print colored("\n\n---------- Re-Checking the Output ----------", "green")
    #print "\n" + str(len(output_bins.keys()))
    #print "\n" + str(output_bins['365day']['days'].keys())
    #print "\n" + str(output_bins['365day']['days']['day_1'].keys())
    #print "\n" + str(output_bins['365day']['days']['day_1']['rooms'].items())
    #print "\n" + str(output_bins['365day']['days']['day_1']['B'])
    #print "\n" + str(output_bins['365day']['days']['day_1']['rooms']['Room-1'])
    print "\n" + str(output_bins['summer']['days']['day_1']['walls'].items())
    print "\n" + str(output_bins['winter']['days']['day_1']['walls'].items())
    #print "\n" + str(output_bins['365day']['days']['day_1']['walls']['Wall-1'])
    #print "\n" + str(output_bins['correlations']['rooms']['Room-1'])
    #print "\n" + str(len(output_bins['correlations']['rooms'].items()))
    #print "\n" + str(output_bins['correlations']['rooms'].items())

    #print output_bins['winter'].keys()
    #print len(output_bins['winter'])
    #print output_bins['winter']['wall_1'].keys()
    #print len(output_bins['winter']['wall_1'])
    #print output_bins['winter']['wall_1']['H'].keys()
    #print len(output_bins['winter']['wall_1']['H'])
    #print output_bins['summer']['wall_2']['S']['day_1']
    #print output_bins['summer']['wall_2']['T']['day_1']

    #print colored("\n---------- Terminating ---------- \n\n", "green")


def get_filepaths(_dir):
    tmp = OrderedDict()
    for dirpath, _, filenames in os.walk(_dir):
        for f in filenames:
            if f.endswith('.csv'):
                filename = os.path.splitext(f)[0]
                tmp[int(filename)] = os.path.abspath(os.path.join(dirpath, f))
    tmp = map(lambda x: tmp[x], sorted(tmp))  # Sort by filename
    return tmp

# Driver Programme
if __name__ == "__main__":

	wall_info = OrderedDict()
	room_info_type1 = OrderedDict()
	room_info_type2 = OrderedDict()
	wall_param_names = []
	room_param_type1_names = []
	room_param_type2_names = []
	
	input_info_file_path = input("'WALL' AND 'ROOM' INFO FILE PATH (Full path in quotes): ")
	input_dir = input('INPUT FOLDER (Full path in quotes): ')
	output_dir = input('OUTPUT FOLDER (Full path in quotes): ')
	output_batch_size = input('OUTPUT BATCH SIZE: ')
	
	print "\nProcessing..\n"
	print "Reading the Input Wall/Room Information..\n"
	read_input_info(input_info_file_path)	# Read Wall/Room info data from the excel file specified.
	
	#print wall_info
	#print wall_param_names
	#print room_info_type1
	#print room_info_type2
	#print room_param_type1_names
	#print room_param_type2_names
	
	
	if not os.path.exists(output_dir):
		print "Creating Output directory.."
		os.makedirs(output_dir)			# Create output directory if not existing
		
	os.chdir(output_dir)   # Change the current directory of program to the specified output_dir

	filepaths = get_filepaths(input_dir)  # Get all files recursively from the input directory

	output_batch_size = int(output_batch_size)  # Output batch size

	input_paths = []
	for i in xrange(0, len(filepaths), output_batch_size):
		input_paths.append(filepaths[i:i+output_batch_size])

	#day_no = input('Input "day number" to process: ')
	#for room_name_old, room_name_new in room_name_no_map.items():
	#  prepare_task7(room_name_old, room_name_new, day_no)


	for i, batch in enumerate(input_paths):
		print "\nStarting Batch-"+str(i+1)+" .."
		batch_filename = 'batch_'+str(i+1)+'.xlsx'
		batch_file_path = os.path.join(output_dir, batch_filename)
		
		#if not os.path.exists(batch_dir_path):
		#	os.makedirs(batch_dir_path)			# Create output batch directory if not existing
			
		print "Preparing a new Excel Workbook for this output batch.."
		wb = openpyxl.Workbook()
		wb = prepare_task1(wb)
		wb = prepare_task2(wb)
		wb = prepare_task3(wb)
		wb = prepare_task4(wb)
		wb = prepare_task5(wb)
		#wb.save(batch_filename)

		for filepath in batch:
			complete_filename = os.path.basename(filepath)
			filename = os.path.splitext(complete_filename)[0]

			print ">> Processing " +  complete_filename

			# Global Variables
			col_map = {}
			col_name_map = {}
			col_names = []
			data_bins = {}
			output_bins = {}

			# Methods
			data = read_csvfile(filepath)
			create_data_bins(data)
			create_data_groups()

			# Check Data
			# recheck_data()
			# check_output()

			# Tasks
			do_task1()
			do_task2()
			do_task3()
			do_task4()
			do_task5()
			#do_task6()
			#do_task7(day_no)

			# Export data to excel file output
			wb = export_task1(wb, filename)
			wb = export_task2(wb, filename)
			wb = export_task3(wb, filename)
			wb = export_task4(wb, filename)
			wb = export_task5(wb, filename)
			wb.save(batch_filename)
			
			#for room_name_old, room_name_new in room_name_no_map.items():
			# export_task7(room_name_old, room_name_new, day_no)

	print "\nFinishing.. Done.\n"
	# End of the script
